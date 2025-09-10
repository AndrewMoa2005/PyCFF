# -*- coding: utf-8 -*-

# reference: https://blog.csdn.net/weixin_40134371/article/details/138843448

import re
import openpyxl
import xlrd, xlwt
import csv
from PySide6.QtCore import Qt, Signal, Slot, qDebug
from PySide6.QtGui import QAction, QKeySequence, QShortcut, QColor
from PySide6.QtWidgets import (
    QTableWidget,
    QTableWidgetSelectionRange,
    QDialog,
    QVBoxLayout,
    QApplication,
    QMenu,
    QMessageBox,
    QGroupBox,
    QRadioButton,
    QPushButton,
    QHBoxLayout,
    QLabel,
    QSpinBox,
    QCheckBox,
    QInputDialog,
    QLineEdit,
    QListWidget,
    QDialogButtonBox,
)

from .clevertwitem import CleverTableWidgetItem as CTWItem


class CleverTableWidget(QTableWidget):
    ContentsChangeSignal = Signal(int, int)
    ItemSelectionSignal = Signal()

    # 构造函数
    def __init__(
        self,
        parent=None,
        rowCount=0,
        columnCount=0,
        editable=True,
        shortcut=True,
        infinite=False,
    ):
        """
        智能表格控件
        Args:
            parent (QWidget, 可选): 父控件
            rowCount (int, 可选): 行数. 默认 0.
            columnCount (int, 可选): 列数. 默认 0.
            editable (bool, 可选): 是否可编辑. 默认 True.
            shortcut (bool, 可选): 是否启用快捷键. 默认 True.
            infinite (bool, 可选): 是否无限行数. 默认 False.
        """
        super().__init__(parent, rowCount=rowCount, columnCount=columnCount)
        self.editable = editable
        self.infinite = infinite
        self.shortcut = shortcut
        self.copy_action = QAction(self.tr("复制"), self)
        self.paste_action = QAction(self.tr("粘贴"), self)
        self.clear_action = QAction(self.tr("清空"), self)
        self.delete_action = QAction(self.tr("删除"), self)
        self.align_left_action = QAction(self.tr("左对齐"), self)
        self.align_right_action = QAction(self.tr("右对齐"), self)
        self.align_center_action = QAction(self.tr("居中对齐"), self)
        self.highlight_action = QAction(self.tr("高亮"), self)
        self.cancle_highlight_action = QAction(self.tr("取消高亮"), self)
        self.bold_action = QAction(self.tr("加粗"), self)
        self.cancle_bold_action = QAction(self.tr("取消加粗"), self)
        self.copy_action.triggered.connect(self.copy_selection)
        self.paste_action.triggered.connect(self.paste_selection)
        self.clear_action.triggered.connect(self.clear_selection)
        self.delete_action.triggered.connect(self.delete_selection)
        self.align_left_action.triggered.connect(self.align_left)
        self.align_right_action.triggered.connect(self.align_right)
        self.align_center_action.triggered.connect(self.align_center)
        self.highlight_action.triggered.connect(self.highlight_background(choice="Y"))
        self.cancle_highlight_action.triggered.connect(
            self.highlight_background(choice="T")
        )
        self.bold_action.triggered.connect(self.bold_text(choice="Y"))
        self.cancle_bold_action.triggered.connect(self.bold_text(choice="N"))
        # 移动操作
        self.move_up_action = QAction(self.tr("上移"), self)
        self.move_down_action = QAction(self.tr("下移"), self)
        self.move_left_action = QAction(self.tr("左移"), self)
        self.move_right_action = QAction(self.tr("右移"), self)
        self.move_up_action.triggered.connect(self.move_up)
        self.move_down_action.triggered.connect(self.move_down)
        self.move_left_action.triggered.connect(self.move_left)
        self.move_right_action.triggered.connect(self.move_right)
        self.move_up_action.setShortcut("Alt+Up")
        self.move_down_action.setShortcut("Alt+Down")
        self.move_left_action.setShortcut("Alt+Left")
        self.move_right_action.setShortcut("Alt+Right")
        # 排序、替换相关操作
        self.ascending_col_action = QAction(self.tr("列升序"), self)
        self.descending_col_action = QAction(self.tr("列降序"), self)
        self.reverse_col_action = QAction(self.tr("列逆序"), self)
        self.paste_col_action = QAction(self.tr("粘贴替换列"), self)
        self.float_col_action = QAction(self.tr("上浮数字"), self)
        self.ascending_col_action.triggered.connect(self.sort_col_ascending)
        self.descending_col_action.triggered.connect(self.sort_col_descending)
        self.reverse_col_action.triggered.connect(self.reverse_col)
        self.paste_col_action.triggered.connect(self.paste_col)
        self.float_col_action.triggered.connect(self.float_col)
        self.ascending_col_action.setShortcut("Alt+A")
        self.descending_col_action.setShortcut("Alt+D")
        self.reverse_col_action.setShortcut("Alt+R")
        self.paste_col_action.setShortcut("Alt+V")
        self.float_col_action.setShortcut("Alt+U")
        # 设置浮点数的显示格式
        self.digital_format_action = QAction(self.tr("设置数字格式"), self)
        self.digital_format_action.triggered.connect(self.digital_format)
        # 列表转置
        self.transpose_action = QAction(self.tr("列表转置"), self)
        self.transpose_action.triggered.connect(self.transpose_table)
        # 设置快捷键
        self.copy_action.setShortcut("Ctrl+C")
        self.paste_action.setShortcut("Ctrl+V")
        self.clear_action.setShortcut("Ctrl+0")
        self.delete_action.setShortcut("Del")
        self.align_left_action.setShortcut("Ctrl+L")
        self.align_right_action.setShortcut("Ctrl+R")
        self.align_center_action.setShortcut("Ctrl+E")
        # 单元格右键菜单
        self.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.customContextMenuRequested.connect(self.showContextMenu)
        # 列标题栏右键菜单
        self.hh = self.horizontalHeader()
        self.hh.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.hh.customContextMenuRequested.connect(self.showContextMenu)
        self.selected_col = None
        # 行标题栏右键菜单
        self.vh = self.verticalHeader()
        self.vh.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.vh.customContextMenuRequested.connect(self.showContextMenu)
        self.selected_row = None
        if not self.editable:
            self.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.shortcut_init = False  # 快捷键初始化标识
        if self.shortcut:
            self.enable_shortcut()
        # 内容所覆盖的最大坐标，以0为起始
        self.max_content_row = 0
        self.max_content_col = 0
        self.cellChanged.connect(self._cell_change_)
        self.whole_rows = []
        self.whole_cols = []
        self.selected_corner_cols = []
        self.selected_corner_rows = []
        self.itemSelectionChanged.connect(self._item_selection_changed)
        self.cellDoubleClicked.connect(self._cell_double_clicked_)

    def _cell_double_clicked_(self, row: int, col: int):
        """
        单元格双击时触发
        """
        item = self.item(row, col)
        text = item.text() if item is not None else ""
        if item is None or not isinstance(item, CTWItem):
            self.setItem(row, col, CTWItem(""))
            item = self.item(row, col)
            item.setText(text)
            qDebug(f"Replace item {row, col} with CTWItem")
        display_text = item.data(Qt.ItemDataRole.DisplayRole)
        if display_text != text:
            item.setData(Qt.ItemDataRole.DisplayRole, text)

    def _cell_change_(self, row: int, col: int):
        """
        单元格内容改变时触发，将默认的QTableWidgetItem替换为CleverTableWidgetItem
        """
        item = self.item(row, col)
        if item is None:
            return
        text = item.text() if item.text() else ""
        if not isinstance(item, CTWItem):
            self.setItem(row, col, CTWItem(""))
            item = self.item(row, col)
            item.setText(text)
            qDebug(f"Replace item {row, col} with CTWItem")
        self._max_content_pos()

    def get_item_from_label(self, label: str) -> CTWItem | None:
        """
        根据Excel风格的单元格标签获取对应的单元格对象
        Args:
            label (str): Excel风格的单元格标签，如"A1", "B2"
        Returns:
            CTWItem | None: 对应的单元格对象，如果标签无效或单元格不存在则返回None
        """
        match = re.match(r"([A-Za-z]+)(\d+)", label)
        if not match:
            return None
        col_str, row_str = match.groups()
        col = 0
        for char in col_str.upper():
            col = col * 26 + (ord(char) - ord("A") + 1)
        col -= 1  # 转换为0基索引
        row = int(row_str) - 1  # 转换为0基索引
        if 0 <= col < self.columnCount() and 0 <= row < self.rowCount():
            return self.item(row, col)
        return None

    def get_label_from_item(self, item: CTWItem) -> str | None:
        """
        根据单元格对象获取对应的Excel风格的单元格标签
        Args:
            item (CTWItem): 单元格对象
        Returns:
            str | None: 对应的Excel风格的单元格标签，如果单元格对象无效则返回None
        """
        if item is None:
            return None
        row = item.row()
        col = item.column()
        if row < 0 or col < 0:
            return None
        col_str = ""
        while col >= 0:
            col_str = chr(col % 26 + ord("A")) + col_str
            col = col // 26 - 1
        return f"{col_str}{row + 1}"

    def _item_selection_changed(self):
        self.whole_cols = self._is_whole_column_selected()
        qDebug(f"Selected whole columns: {self.whole_cols}")
        self.whole_rows = self._is_whole_row_selected()
        qDebug(f"Selected whole rows: {self.whole_rows}")
        self._get_selected_items_right_down_corner()
        qDebug(
            f"Selected items right down corner: {self.selected_corner_cols, self.selected_corner_rows}"
        )
        self._max_content_pos()
        self.ItemSelectionSignal.emit()

    def resizeEvent(self, event):
        # 窗口大小调整时，自动填充可见区域
        if self.infinite:
            self._fill_visible_area()
        super().resizeEvent(event)

    def initial_shortcut(
        self, key: str, func: callable, context=Qt.WidgetShortcut
    ) -> QShortcut:
        """
        初始化快捷键
        Args:
            key (str): 快捷键
            func (callable): 快捷键触发的函数
            context (Qt.WidgetShortcut, 可选): 快捷键的上下文. 默认为 Qt.WidgetShortcut.
        Returns:
            QShortcut: 初始化后的快捷键
        """
        s = QShortcut(QKeySequence(key), self)
        s.activated.connect(func)
        s.setContext(context)
        return s

    def enable_shortcut(self):
        """
        启用快捷键
        """
        if self.shortcut is False:
            self.shortcut = True
        # 初始化
        if self.shortcut_init is False:
            # 移动操作快捷键
            self.short_a_up = self.initial_shortcut("Alt+Up", self.move_up)
            self.short_a_down = self.initial_shortcut("Alt+Down", self.move_down)
            self.short_a_left = self.initial_shortcut("Alt+Left", self.move_left)
            self.short_a_right = self.initial_shortcut("Alt+Right", self.move_right)
            # 排序、替换相关操作快捷键
            self.short_a_a = self.initial_shortcut("Alt+A", self.sort_col_ascending)
            self.short_a_d = self.initial_shortcut("Alt+D", self.sort_col_descending)
            self.short_a_r = self.initial_shortcut("Alt+R", self.reverse_col)
            self.short_a_v = self.initial_shortcut("Alt+V", self.paste_col)
            self.short_a_u = self.initial_shortcut("Alt+U", self.float_col)
            # 复制、对齐操作相关快捷键
            self.short_c_c = self.initial_shortcut("Ctrl+C", self.copy_selection)
            self.short_c_v = self.initial_shortcut("Ctrl+V", self.paste_selection)
            self.short_c_0 = self.initial_shortcut("Ctrl+0", self.clear_selection)
            self.short_del = self.initial_shortcut("Del", self.delete_selection)
            self.short_c_l = self.initial_shortcut("Ctrl+L", self.align_left)
            self.short_c_r = self.initial_shortcut("Ctrl+R", self.align_right)
            self.short_c_e = self.initial_shortcut("Ctrl+E", self.align_center)
            # 设置标签
            self.shortcut_init = True
        else:
            # 移动操作快捷键
            self.short_a_up.setEnabled(True)
            self.short_a_down.setEnabled(True)
            self.short_a_left.setEnabled(True)
            self.short_a_right.setEnabled(True)
            # 排序、替换相关操作快捷键
            self.short_a_a.setEnabled(True)
            self.short_a_d.setEnabled(True)
            self.short_a_r.setEnabled(True)
            self.short_a_v.setEnabled(True)
            self.short_a_u.setEnabled(True)
            # 复制、对齐操作相关快捷键
            self.short_c_c.setEnabled(True)
            self.short_c_v.setEnabled(True)
            self.short_c_0.setEnabled(True)
            self.short_del.setEnabled(True)
            self.short_c_l.setEnabled(True)
            self.short_c_r.setEnabled(True)
            self.short_c_e.setEnabled(True)

    def disable_shortcut(self):
        """
        禁用快捷键
        """
        if self.shortcut is True:
            self.shortcut = False
        if self.shortcut_init is True:
            # 移动操作快捷键
            self.short_a_up.setEnabled(False)
            self.short_a_down.setEnabled(False)
            self.short_a_left.setEnabled(False)
            self.short_a_right.setEnabled(False)
            # 排序、替换相关操作快捷键
            self.short_a_a.setEnabled(False)
            self.short_a_d.setEnabled(False)
            self.short_a_r.setEnabled(False)
            self.short_a_v.setEnabled(False)
            self.short_a_u.setEnabled(False)
            # 复制、对齐操作相关快捷键
            self.short_c_c.setEnabled(False)
            self.short_c_v.setEnabled(False)
            self.short_c_0.setEnabled(False)
            self.short_del.setEnabled(False)
            self.short_c_l.setEnabled(False)
            self.short_c_r.setEnabled(False)
            self.short_c_e.setEnabled(False)

    def is_shortcut(self):
        """
        判断是否启用快捷键
        """
        return self.shortcut

    def is_infinite(self):
        """
        判断是否启用无限行数
        """
        return self.infinite

    def is_editable(self):
        """
        判断是否启用可编辑
        """
        return self.editable

    def enable_infinite(self):
        """
        启用无限行数
        """
        self.infinite = True
        # 连接滚动条信号
        self.horizontalScrollBar().valueChanged.connect(self._handle_horizontal_scroll)
        self.verticalScrollBar().valueChanged.connect(self._handle_vertical_scroll)
        # 初始填充可见区域
        self._fill_visible_area()

    def _handle_horizontal_scroll(self, value):
        """处理水平滚动"""
        scrollbar = self.horizontalScrollBar()
        if value == scrollbar.maximum():
            # 滚动到最右侧，添加新列
            self._add_columns_at_end()
        else:
            # 回滚，检查并删除空列
            self._remove_columns_at_end()
        self.renumber_header_col()

    def _handle_vertical_scroll(self, value):
        """处理垂直滚动"""
        scrollbar = self.verticalScrollBar()
        if value == scrollbar.maximum():
            # 滚动到底部，添加新行
            self._add_rows_at_end()
        else:
            # 回滚，检查并删除空行
            self._remove_rows_at_end()
        self.renumber_header_row()

    def _fill_visible_area(self):
        """填充当前可见区域"""
        viewport = self.viewport()
        # 获取viewport的右下角坐标(x0, y0)
        x0 = viewport.width()
        y0 = viewport.height()
        # 获取最后一个单元格的右下角坐标(x1, y1)
        last_row = self.rowCount() - 1
        last_col = self.columnCount() - 1
        x1 = self.columnViewportPosition(last_col) + self.columnWidth(last_col)
        y1 = self.rowViewportPosition(last_row) + self.rowHeight(last_row)
        self.renumber_header()
        # 当最后一个单元格的坐标小于viewport坐标时，添加新行或新列
        while x1 < x0:
            self._add_columns_at_end()
            last_col = self.columnCount() - 1
            x1 = self.columnViewportPosition(last_col) + self.columnWidth(last_col)
        while y1 < y0:
            self._add_rows_at_end()
            last_row = self.rowCount() - 1
            y1 = self.rowViewportPosition(last_row) + self.rowHeight(last_row)
        # 获取最后一个单元格的左上角坐标(x10, y10)
        x10 = self.columnViewportPosition(last_col)
        y10 = self.rowViewportPosition(last_row)
        # 当最后一个单元格的坐标大于viewport坐标时，移除不可见空行或空列
        while x10 > x0:
            if self.col_is_empty(last_col):
                self._remove_columns_at_end()
                last_col = self.columnCount() - 1
                x10 = self.columnViewportPosition(last_col)
            else:
                break
        while y10 > y0:
            if self.row_is_empty(last_row):
                self._remove_rows_at_end()
                last_row = self.rowCount() - 1
                y10 = self.rowViewportPosition(last_row)
            else:
                break

    def _add_rows_at_end(self):
        """在末尾添加新行"""
        current_count = self.rowCount()
        self.setRowCount(current_count + 1)
        qDebug(f"Added new row {current_count} at end")

    def _remove_rows_at_end(self):
        """移除末尾不可见空行"""
        if self.rowCount() <= 0:
            return
        rows = self.rowCount() - 1
        if rows >= 0 and self.row_is_empty(rows):
            if self.whole_rows is not None and rows in self.whole_rows:
                return
            elif (
                self.selected_corner_rows is not None
                and rows in self.selected_corner_rows
            ):
                return
            self.removeRow(rows)
            qDebug(f"Removed empty row {rows} at end")

    def _add_columns_at_end(self):
        """在末尾添加新列"""
        current_count = self.columnCount()
        self.setColumnCount(current_count + 1)
        qDebug(f"Added new column {current_count} at end")

    def _remove_columns_at_end(self):
        """移除末尾不可见空列"""
        if self.columnCount() <= 0:
            return
        cols = self.columnCount() - 1
        if cols >= 0 and self.col_is_empty(cols):
            if self.whole_cols is not None and cols in self.whole_cols:
                return
            elif (
                self.selected_corner_cols is not None
                and cols in self.selected_corner_cols
            ):
                return
            self.removeColumn(cols)
            qDebug(f"Removed empty column {cols} at end")

    def disable_infinite(self):
        """
        禁用无限行数
        """
        self.infinite = False
        # 断开连接滚动条信号
        self.horizontalScrollBar().valueChanged.disconnect(
            self._handle_horizontal_scroll
        )
        self.verticalScrollBar().valueChanged.disconnect(self._handle_vertical_scroll)
        # 清除空行
        self.clear_empty_space()

    def disable_editing(self):
        """
        禁用可编辑
        """
        self.editable = False
        self.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)

    def enable_editing(self):
        """
        启用可编辑
        """
        self.editable = True
        self.setEditTriggers(QTableWidget.EditTrigger.DoubleClicked)

    def change_horizontal_header(self, index):
        """
        编辑列标题
        """
        if self.editable is False:
            return
        item = self.horizontalHeaderItem(index)
        if item is None:
            val = self.model().headerData(index, Qt.Orientation.Horizontal)
            item = CTWItem(str(val))
            self.setHorizontalHeaderItem(index, item)
        old_header = item.text()
        new_header, ok_pressed = QInputDialog.getText(
            self,
            self.tr("设置列标题"),
            self.tr("请输入第%s列标题") % (index + 1),
            QLineEdit.EchoMode.Normal,
            old_header,
        )
        if ok_pressed:
            item.setText(new_header)

    def change_vertical_header(self, index):
        """
        编辑行标题
        """
        if self.editable is False:
            return
        item = self.verticalHeaderItem(index)
        if item is None:
            val = self.model().headerData(index, Qt.Orientation.Vertical)
            item = CTWItem(str(val))
            self.setVerticalHeaderItem(index, item)
        old_header = item.text()
        new_header, ok_pressed = QInputDialog.getText(
            self,
            self.tr("设置行标题"),
            self.tr("请输入第%s行标题") % (index + 1),
            QLineEdit.EchoMode.Normal,
            old_header,
        )
        if ok_pressed:
            item.setText(new_header)

    def _get_selected_items_right_down_corner(self):
        """
        获取选中区域的右下角坐标
        """
        selected_ranges = self.selectedRanges()
        if not selected_ranges:
            return None
        corner_x = []
        corner_y = []
        for selected in selected_ranges:
            corner_x.append(selected.rightColumn())
            corner_y.append(selected.bottomRow())
        self.selected_corner_cols = corner_x
        self.selected_corner_rows = corner_y

    def _is_whole_row_selected(self):
        """
        判断是否选择了整行，返回选中的行号
        """
        selected_ranges = self.selectedRanges()
        if not selected_ranges:
            return None
        row_list = []
        for selected in selected_ranges:
            if (
                selected.rightColumn() - selected.leftColumn() + 1
            ) == self.columnCount():
                for row in range(selected.topRow(), selected.bottomRow() + 1):
                    if row not in row_list:
                        row_list.append(row)
        if len(row_list) > 0:
            return row_list
        return None

    def _is_whole_column_selected(self):
        """
        判断是否选择了整列，返回选中的列号
        """
        selected_ranges = self.selectedRanges()
        if not selected_ranges:
            return None
        col_list = self.whole_rows = []
        for selected in selected_ranges:
            if (selected.bottomRow() - selected.topRow() + 1) == self.rowCount():
                for col in range(selected.leftColumn(), selected.rightColumn() + 1):
                    if col not in col_list:
                        col_list.append(col)
        if len(col_list) > 0:
            return col_list
        return None

    def showContextMenu(self, pos):
        """
        显示上下文菜单
        """
        if (
            pos.x() > self.hh.x()
            and pos.x() < self.hh.x() + self.hh.width()
            and pos.y() > self.hh.y()
            and pos.y() < self.hh.y() + self.hh.height()
        ):
            # 单击点在列标题栏内
            self.selected_col = self.hh.logicalIndexAt(pos)
            self.selectColumn(self.selected_col)
        else:
            self.selected_col = None
        if (
            pos.x() > self.vh.x()
            and pos.x() < self.vh.x() + self.vh.width()
            and pos.y() > self.vh.y()
            and pos.y() < self.vh.y() + self.vh.height()
        ):
            # 单击点在行标题栏内
            self.selected_row = self.vh.logicalIndexAt(pos)
            self.selectRow(self.selected_row)
        else:
            self.selected_row = None
        menu = QMenu(self)
        align_menu = QMenu(self.tr("对齐方式"), self)
        insert_action = QAction(self.tr("插入"), self)
        insert_action_common = QAction(self.tr("插入"), self)
        insert_action_common.triggered.connect(self.insert_base_on_selection)
        menu.addAction(self.copy_action)
        if self.editable:
            menu.addAction(self.paste_action)
            menu.addAction(self.clear_action)
            menu.addAction(self.delete_action)
            align_menu.addAction(self.align_center_action)
        align_menu.addAction(self.align_left_action)
        align_menu.addAction(self.align_right_action)
        if self.editable:
            # 移动操作
            move_menu = QMenu(self.tr("移动单元格"), self)
            move_menu.addAction(self.move_up_action)
            move_menu.addAction(self.move_down_action)
            move_menu.addAction(self.move_left_action)
            move_menu.addAction(self.move_right_action)
            menu.addMenu(move_menu)
            # 插入操作
            selected = self.selectedRanges()
            if len(selected) == 0:
                return
            elif len(selected) > 1:  # 不支持多重选择场景
                return
            else:
                s = selected[0]
                srn = s.bottomRow() - s.topRow() + 1
                scn = s.rightColumn() - s.leftColumn() + 1
                if (
                    srn == self.rowCount()
                ):  # 选中区域行数等于表格最大行数, 选中整列, 向左侧插入新列
                    insert_action.triggered.connect(
                        self.insert_whole_base_on_selection(insert_type="C")
                    )
                    menu.addAction(insert_action)
                elif (
                    scn == self.columnCount()
                ):  # 选中区域列数等于表格最大列数, 选中整行, 向上方插入新行
                    insert_action.triggered.connect(
                        self.insert_whole_base_on_selection(insert_type="R")
                    )
                    menu.addAction(insert_action)
                else:
                    menu.addAction(insert_action_common)
        if self.editable and self.selected_col is not None:
            menu.addSeparator()
            menu.addAction(self.ascending_col_action)
            menu.addAction(self.descending_col_action)
            menu.addAction(self.reverse_col_action)
            menu.addAction(self.paste_col_action)
            menu.addAction(self.float_col_action)
        menu.addSeparator()
        menu.addMenu(align_menu)
        menu.addAction(self.highlight_action)
        menu.addAction(self.bold_action)
        menu.addAction(self.cancle_highlight_action)
        menu.addAction(self.cancle_bold_action)
        menu.addSeparator()
        if self.editable:
            menu.addAction(self.digital_format_action)
        menu.addAction(self.transpose_action)
        menu.exec(self.viewport().mapToGlobal(pos))

    def move_up(self):
        """
        向上移动选中区域
        """
        if self.editable is False:
            return
        selected_ranges = self.selectedRanges()
        if not selected_ranges:
            return
        if len(selected_ranges) > 1:
            return
        top = selected_ranges[0].topRow()
        if top == 0:
            QMessageBox.warning(
                self,
                self.tr("警告"),
                self.tr("已到达顶部，无法向上移动"),
            )
            return
        left = selected_ranges[0].leftColumn()
        right = selected_ranges[0].rightColumn()
        bottom = selected_ranges[0].bottomRow()
        for col in range(left, right + 1):
            for row in range(top, bottom + 1):
                item = self.takeItem(row, col)
                if not item:
                    item = CTWItem().setText("")
                item_prev = self.takeItem(row - 1, col)
                if not item_prev:
                    item_prev = CTWItem().setText("")
                self.setItem(row - 1, col, item)
                self.setItem(row, col, item_prev)
        self.clearSelection()
        self.setRangeSelected(
            QTableWidgetSelectionRange(top - 1, left, bottom - 1, right), True
        )
        self._max_content_pos()

    def move_down(self):
        """
        向下移动选中区域
        """
        if self.editable is False:
            return
        selected_ranges = self.selectedRanges()
        if not selected_ranges:
            return
        if len(selected_ranges) > 1:
            return
        bottom = selected_ranges[0].bottomRow()
        if bottom == self.rowCount() - 1:
            self.insertRow(self.rowCount())
        top = selected_ranges[0].topRow()
        left = selected_ranges[0].leftColumn()
        right = selected_ranges[0].rightColumn()
        for col in range(left, right + 1):
            for row in range(bottom, top - 1, -1):
                item = self.takeItem(row, col)
                if not item:
                    item = CTWItem().setText("")
                item_next = self.takeItem(row + 1, col)
                if not item_next:
                    item_next = CTWItem().setText("")
                self.setItem(row + 1, col, item)
                self.setItem(row, col, item_next)
        self.clearSelection()
        self.setRangeSelected(
            QTableWidgetSelectionRange(top + 1, left, bottom + 1, right), True
        )
        self._max_content_pos()

    def move_left(self):
        """
        向左移动选中区域
        """
        if self.editable is False:
            return
        selected_ranges = self.selectedRanges()
        if not selected_ranges:
            return
        if len(selected_ranges) > 1:
            return
        left = selected_ranges[0].leftColumn()
        if left == 0:
            QMessageBox.warning(
                self,
                self.tr("警告"),
                self.tr("已到达最左侧，无法向左移动"),
            )
            return
        top = selected_ranges[0].topRow()
        right = selected_ranges[0].rightColumn()
        bottom = selected_ranges[0].bottomRow()
        for row in range(top, bottom + 1):
            for col in range(left, right + 1):
                item = self.takeItem(row, col)
                if not item:
                    item = CTWItem().setText("")
                item_prev = self.takeItem(row, col - 1)
                if not item_prev:
                    item_prev = CTWItem().setText("")
                self.setItem(row, col - 1, item)
                self.setItem(row, col, item_prev)
        self.clearSelection()
        self.setRangeSelected(
            QTableWidgetSelectionRange(top, left - 1, bottom, right - 1), True
        )
        self._max_content_pos()

    def move_right(self):
        """
        向右移动选中区域
        """
        if self.editable is False:
            return
        selected_ranges = self.selectedRanges()
        if not selected_ranges:
            return
        if len(selected_ranges) > 1:
            return
        right = selected_ranges[0].rightColumn()
        if right == self.columnCount() - 1:
            self.insertColumn(self.columnCount())
        top = selected_ranges[0].topRow()
        left = selected_ranges[0].leftColumn()
        bottom = selected_ranges[0].bottomRow()
        for row in range(top, bottom + 1):
            for col in range(right, left - 1, -1):
                item = self.takeItem(row, col)
                if not item:
                    item = CTWItem().setText("")
                item_next = self.takeItem(row, col + 1)
                if not item_next:
                    item_next = CTWItem().setText("")
                self.setItem(row, col + 1, item)
                self.setItem(row, col, item_next)
        self.clearSelection()
        self.setRangeSelected(
            QTableWidgetSelectionRange(top, left + 1, bottom, right + 1), True
        )
        self._max_content_pos()

    def digital_format(self):
        """
        在弹出对话框设置选中单元格的数字格式
        """

        selected_ranges = self.selectedRanges()
        if not selected_ranges:
            return
        selected = selected_ranges[0]
        r = selected.topRow()
        c = selected.leftColumn()
        i = self.item(r, c)
        if i is None:
            i_sci = False
            i_dec = None
        else:
            i_sci = i.sci
            i_dec = i.dec

        class FormatDialog(QDialog):
            def __init__(self, parent=None, sci: bool = False, dec: int = None):
                """
                数字格式对话框
                Args:
                    parent (QWidget, optional): 父窗口. Defaults to None.
                    sci (bool, optional): 是否使用科学计数法. Defaults to False.
                    dec (int, optional): 小数点位数. Defaults to None.
                """
                super().__init__(parent)
                self.setWindowTitle(self.tr("设置数字格式"))
                self.resize(300, 150)
                layout = QVBoxLayout(self)
                # 小数点位数
                hbox1 = QHBoxLayout()
                hbox1.addWidget(QLabel(self.tr("小数点位数:")))
                self.decimal_spin = QSpinBox()
                self.decimal_spin.setRange(0, 15)
                self.decimal_spin.setValue(dec if dec is not None else 2)
                hbox1.addWidget(self.decimal_spin)
                layout.addLayout(hbox1)
                # 科学计数法
                hbox2 = QHBoxLayout()
                self.sci_checkbox = QCheckBox(self.tr("使用科学计数法"))
                self.sci_checkbox.setChecked(sci)
                hbox2.addWidget(self.sci_checkbox)
                layout.addLayout(hbox2)
                # 按钮
                btn_box = QHBoxLayout()
                self.clear_btn = QPushButton(self.tr("清除格式"))
                self.ok_btn = QPushButton(self.tr("确定"))
                self.cancel_btn = QPushButton(self.tr("取消"))
                btn_box.addWidget(self.clear_btn)
                btn_box.addWidget(self.ok_btn)
                btn_box.addWidget(self.cancel_btn)
                layout.addLayout(btn_box)
                self.ok_btn.clicked.connect(self.accept)
                self.cancel_btn.clicked.connect(self.reject)
                self.clear_btn.clicked.connect(self.clear_format)
                self.clear_label = False

            def clear_format(self):
                self.clear_label = True
                qDebug("clear format")
                super().accept()

            def accept(self):
                self.clear_label = False
                qDebug(
                    "set format: decimals = {}, use_sci = {}".format(
                        self.decimal_spin.value(), self.sci_checkbox.isChecked()
                    )
                )
                super().accept()

        dlg = FormatDialog(self, i_sci, i_dec)
        if dlg.exec() != QDialog.Accepted:
            return
        should_clear = dlg.clear_label
        decimals = dlg.decimal_spin.value()
        use_sci = dlg.sci_checkbox.isChecked()
        selected = self.selectedRanges()
        top_row = selected[0].topRow()
        bottom_row = selected[0].bottomRow()
        left_column = selected[0].leftColumn()
        right_column = selected[0].rightColumn()
        for row in range(top_row, bottom_row + 1):
            for column in range(left_column, right_column + 1):
                item = self.item(row, column)
                if item is not None:
                    if not isinstance(item, CTWItem):
                        item = CTWItem(item.text())
                        self.setItem(row, column, item)
                    if should_clear:
                        item.clearFormat()
                    else:
                        item.setFormat(decimals, use_sci)

    def get_selected_columns_list(self):
        """
        获取选中的列索引列表
        """
        selected_items = self.selectedIndexes()
        if not selected_items:
            return []
        selected_list = []
        for i in selected_items:
            if i.column() < 0 or i.column() >= self.columnCount():
                continue
            selected_list.append(i.column())
        # remove duplicate elements
        selected_list = sorted(list(set(selected_list)))
        return selected_list

    def sort_col_ascending(self):
        """
        按升序对选中的列进行排序
        """
        if self.editable is False:
            return
        selected_list = self.get_selected_columns_list()
        if not selected_list:
            return
        for selected_column in selected_list:
            items = []
            for row in range(self.rowCount()):
                item = self.item(row, selected_column)
                if item:
                    items.append(item)

            def get_numeric_value(item):
                try:
                    return float(item.text())
                except ValueError:
                    return float("inf")

            items.sort(key=get_numeric_value)
            for row in range(self.rowCount()):
                self.takeItem(row, selected_column)
            for new_row, item in enumerate(items):
                self.setItem(new_row, selected_column, item)
        self._max_content_pos()

    def sort_col_descending(self):
        """
        按降序对选中的列进行排序
        """
        if self.editable is False:
            return
        self.sort_col_ascending()
        self.reverse_col()

    def reverse_col(self):
        """
        反转选中的列
        """
        if self.editable is False:
            return
        selected_list = self.get_selected_columns_list()
        if not selected_list:
            return
        for selected_column in selected_list:
            items = []
            for row in range(self.rowCount()):
                item = self.item(row, selected_column)
                if item:
                    items.append(item)
            items = items[::-1]
            for row in range(self.rowCount()):
                self.takeItem(row, selected_column)
            for new_row, item in enumerate(items):
                self.setItem(new_row, selected_column, item)
        self._max_content_pos()

    def filter_string_to_float_list(self, input_str: str) -> list:
        """
        从字符串中提取浮点数列表
        """
        pattern = r"[-+]?\d+\.?\d*(?:[eE][-+]?\d+)?"
        matches = re.findall(pattern, input_str)
        output_list = []
        for num_str in matches:
            try:
                output_list.append(float(num_str))
            except ValueError:
                continue
        return output_list

    def paste_col(self):
        """
        粘贴字符串中的数字到所选列，以任意非数字字符分隔
        """
        if self.editable is False:
            return
        selected_list = self.get_selected_columns_list()
        if len(selected_list) != 1:
            QMessageBox.warning(
                self, self.tr("警告"), self.tr("只能选择一列进行粘贴替换操作")
            )
            return
        str_data = QApplication.clipboard().text()
        float_list = self.filter_string_to_float_list(str_data)
        if not float_list:
            QMessageBox.warning(
                self, self.tr("警告"), self.tr("粘贴数据中不包含有效数字字符")
            )
            return
        if len(float_list) > self.rowCount():
            self.setRowCount(len(float_list))
        col = selected_list[0]
        for row in range(self.rowCount()):
            if row < len(float_list):
                item = CTWItem(str(float_list[row]))
                # item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.setItem(row, col, item)
            else:
                self.setItem(row, col, CTWItem(""))
        self._max_content_pos()

    def float_col(self):
        """
        将选中的列的数字上浮到顶部
        """
        if self.editable is False:
            return
        selected_list = self.get_selected_columns_list()
        if not selected_list:
            return
        for col in selected_list:
            number_list = []
            for row in range(self.rowCount()):
                item = self.item(row, col)
                text = item.text() if item is not None else ""
                try:
                    float(text)
                except ValueError or AttributeError:
                    continue
                number_list.append(text)
            for row in range(self.rowCount()):
                if row < len(number_list):
                    item = CTWItem(str(number_list[row]))
                    self.setItem(row, col, item)
                else:
                    self.setItem(row, col, CTWItem(""))
        self._max_content_pos()

    def row_is_empty(self, row: int) -> bool:
        for col in range(self.columnCount() - 1, 0, -1):
            item = self.item(row, col)
            text = item.text() if item is not None else ""
            if text in ["", " "]:
                continue
            else:
                return False
        return True

    def col_is_empty(self, col: int) -> bool:
        for row in range(self.rowCount() - 1, 0, -1):
            item = self.item(row, col)
            text = item.text() if item is not None else ""
            if text in ["", " "]:
                continue
            else:
                return False
        return True

    def _max_content_pos(self):
        """
        获取最大内容所在位置, 从0开始
        """
        max_row = self.rowCount() - 1
        max_col = self.columnCount() - 1
        for row in range(max_row, 0, -1):
            if self.row_is_empty(row):
                max_row -= 1
            else:
                break
        for col in range(max_col, 0, -1):
            if self.col_is_empty(col):
                max_col -= 1
            else:
                break
        self.max_content_row = max_row
        self.max_content_col = max_col
        self.ContentsChangeSignal.emit(max_row, max_col)

    def get_max_content_pos(self):
        self._max_content_pos()
        return self.max_content_row, self.max_content_col

    def clear_empty_row(self):
        """
        清除空行
        """
        if self.infinite:
            return
        if self.editable is False:
            return
        for row in range(self.rowCount() - 1, -1, -1):
            n = 0
            for clo in range(self.columnCount()):
                item = self.item(row, clo)
                text = item.text() if item is not None else ""
                if text in ["", " "]:
                    n += 1
                if n == self.columnCount():
                    self.removeRow(row)
        """
        while self.max_content_row + 1 < self.rowCount():
            self.removeRow(self.rowCount() - 1)
        """

    def clear_empty_col(self):
        """
        清除空列
        """
        if self.infinite:
            return
        if self.editable is False:
            return
        for clo in range(self.columnCount() - 1, -1, -1):
            n = 0
            for row in range(self.rowCount()):
                item = self.item(row, clo)
                text = item.text() if item is not None else ""
                if text in ["", " "]:
                    n += 1
                if n == self.rowCount():
                    self.removeColumn(clo)
        """
        while self.max_content_col + 1 < self.columnCount():
            self.removeColumn(self.columnCount() - 1)
        """

    def clear_empty_space(self):
        """
        清除表格
        """
        if self.infinite:
            return
        self.clear_empty_col()
        self.clear_empty_row()

    def transpose_table(self):
        """
        列表转置
        """
        self.clear_empty_space()
        if self.rowCount() == 0 or self.columnCount() == 0:
            QMessageBox.warning(
                self, self.tr("警告"), self.tr("表格为空，无法进行转置操作")
            )
            return

        def trans():
            for row in range(self.rowCount()):
                # 交换单元格
                for col in range(row + 1, self.columnCount()):
                    if row != col:
                        item_rc = self.takeItem(row, col)
                        item_cr = self.takeItem(col, row)
                        self.setItem(row, col, item_cr)
                        self.setItem(col, row, item_rc)
                # 交换表头
                if not self.infinite:
                    head_col = self.takeHorizontalHeaderItem(row)
                    head_row = self.takeVerticalHeaderItem(row)
                    self.setHorizontalHeaderItem(row, head_row)
                    self.setVerticalHeaderItem(row, head_col)
                else:
                    self.renumber_header()

        if self.rowCount() == self.columnCount():
            trans()
        elif self.rowCount() > self.columnCount():
            self.setColumnCount(self.rowCount())
            trans()
        else:
            self.setRowCount(self.columnCount())
            trans()
        if not self.editable:
            self.enable_editing()
            self.clear_empty_space()
            self.disable_editing()
        else:
            self.clear_empty_space()

    def renumber_header_col(self):
        """
        重编号列标题为Excel风格的字母顺序
        """
        if self.editable is False:
            return

        def num_to_letters(n):
            letters = ""
            while n > 0:
                n -= 1
                letters = chr(ord("A") + n % 26) + letters
                n = n // 26
            return letters

        for col in range(self.columnCount()):
            header = self.horizontalHeaderItem(col)
            if header is None:
                header = CTWItem()
                self.setHorizontalHeaderItem(col, header)
            header.setText(num_to_letters(col + 1))

    def renumber_header_row(self):
        """
        重编号行标题
        """
        if self.editable is False:
            return
        for row in range(self.rowCount()):
            header = self.verticalHeaderItem(row)
            if header is None:
                header = CTWItem()
                self.setVerticalHeaderItem(row, header)
            header.setText(str(row + 1))

    def renumber_header(self):
        """
        重编号列和行标题
        """
        self.renumber_header_col()
        self.renumber_header_row()

    def get_first_empty_row_id(self):
        """
        获取第一个空行的ID
        """
        first_empty_row_id = 0
        for row in range(self.rowCount()):
            exist_item = False
            for column in range(self.columnCount()):
                if self.item(row, column) is not None and self.item(
                    row, column
                ).text() not in ["", " "]:
                    exist_item = True
                    break
            if not exist_item:
                return first_empty_row_id
            first_empty_row_id += 1
        return first_empty_row_id

    def read_table_context(self):
        """
        读取表格内容
        """
        context = [
            [
                self.item(row, col).text() if self.item(row, col) is not None else ""
                for col in range(self.columnCount())
            ]
            for row in range(self.rowCount())
        ]
        return context

    def _insert_add_row_helper(self, selected_top, selected_bottom):
        finaly_unempty_row_id = self.rowCount()
        end = False
        for row in sorted(range(self.rowCount()), reverse=True):
            finaly_unempty_row_id -= 1
            for column in range(selected_top, selected_bottom + 1):
                if self.item(row, column) is not None and self.item(
                    row, column
                ).text() not in ["", " "]:
                    end = True
                    break
            if end:
                break
        return finaly_unempty_row_id

    def _insert_add_col_helper(self, selected_left, selected_right):
        finaly_unempty_col_id = self.columnCount()
        end = False
        for column in sorted(range(self.columnCount()), reverse=True):
            finaly_unempty_col_id -= 1
            for row in range(selected_left, selected_right + 1):
                if self.item(row, column) is not None and self.item(
                    row, column
                ).text() not in ["", " "]:
                    end = True
                    break
            if end:
                break
        return finaly_unempty_col_id

    def _judge_rectangular_selected(self):
        selection = self.selectedRanges()
        if len(selection) != 1:
            return False
        return True

    def copy_selection(self):
        """
        复制表格选中内容到剪贴板
        """
        if not self._judge_rectangular_selected():
            return
        selected = self.selectedRanges()[0]
        text = "\n".join(
            [
                "\t".join(
                    [
                        (
                            self.item(row, col).text()
                            if self.item(row, col) is not None
                            else ""
                        )
                        for col in range(
                            selected.leftColumn(), selected.rightColumn() + 1
                        )
                    ]
                )
                for row in range(selected.topRow(), selected.bottomRow() + 1)
            ]
        )
        QApplication.clipboard().setText(text)

    def paste_selection(self):
        """
        从剪贴板粘贴内容到表格
        """
        if self.editable is False:
            return
        if not self._judge_rectangular_selected():
            return
        selected = self.selectedRanges()[0]
        text = QApplication.clipboard().text()
        rows = text.split("\n")
        if "" in rows:
            rows.remove("")

        for r, row in enumerate(rows):
            if selected.topRow() + r >= self.rowCount():
                self.insertRow(selected.topRow() + r)
            cols = row.split("\t")
            for c, text in enumerate(cols):
                if selected.leftColumn() + c >= self.columnCount():
                    self.insertColumn(selected.leftColumn() + c)
                self.setItem(
                    selected.topRow() + r,
                    selected.leftColumn() + c,
                    CTWItem(text),
                )
        self._max_content_pos()

    def clear_selection(self):
        """
        清除选中区域内容
        """
        if self.editable is False:
            return
        for item in self.selectedItems():
            self.setItem(item.row(), item.column(), CTWItem(""))
        self._max_content_pos()

    def delete_selection(self):
        """
        删除选中区域内容
        """
        if self.editable is False:
            return
        if not self._judge_rectangular_selected():
            return
        if self.selected_col is not None:
            self._delete_operations("Delete Selected Cols")
        elif self.selected_row is not None:
            self._delete_operations("Delete Selected Rows")
        elif self.whole_cols is not None:
            self._delete_operations("Delete Selected Cols")
        elif self.whole_rows is not None:
            self._delete_operations("Delete Selected Rows")
        else:
            delete_dialog = DeleteInsertDialog(dialog_type="delete")
            delete_dialog.DelSignal.connect(self._delete_operations)
            delete_dialog.setWindowModality(Qt.WindowModality.ApplicationModal)
            delete_dialog.exec()

    @Slot(str)
    def _delete_operations(self, message):
        selected = self.selectedRanges()[0]
        if message == "Move Left":
            selected_cols_num = selected.rightColumn() - selected.leftColumn() + 1
            start_col_index = selected.leftColumn() + selected_cols_num
            for col in range(start_col_index, self.columnCount() + selected_cols_num):
                ori_col = col - selected_cols_num
                for row in range(selected.topRow(), selected.bottomRow() + 1):
                    if col < self.columnCount():
                        text = (
                            self.item(row, col).text()
                            if self.item(row, col) is not None
                            else ""
                        )
                        self.setItem(row, ori_col, CTWItem(text))
                    else:
                        self.setItem(row, ori_col, CTWItem(""))
        elif message == "Move Up":
            selected_rows_num = selected.bottomRow() - selected.topRow() + 1
            start_row_index = selected.topRow() + selected_rows_num
            for row in range(start_row_index, self.rowCount() + selected_rows_num):
                ori_row = row - selected_rows_num
                for col in range(selected.leftColumn(), selected.rightColumn() + 1):
                    if row < self.rowCount():
                        text = (
                            self.item(row, col).text()
                            if self.item(row, col) is not None
                            else ""
                        )
                        self.setItem(ori_row, col, CTWItem(text))
                    else:
                        self.setItem(ori_row, col, CTWItem(""))
        elif message == "Delete Selected Rows":
            # 从最后一列开始删除，避免删除后索引变化
            for row in sorted(
                range(selected.topRow(), selected.bottomRow() + 1), reverse=True
            ):
                self.removeRow(row)
        elif message == "Delete Selected Cols":
            # 从最后一列开始删除，避免删除后索引变化
            for column in sorted(
                range(selected.leftColumn(), selected.rightColumn() + 1), reverse=True
            ):
                self.removeColumn(column)
        else:
            qDebug("Empty Message")
        self.selected_col = None
        self.selected_row = None
        self.renumber_header()
        self._max_content_pos()

    def insert_whole_base_on_selection(self, insert_type):
        def insert():
            selected = self.selectedRanges()[0]
            if insert_type == "R":
                row_id = selected.topRow()
                for i in range(selected.bottomRow() - selected.topRow() + 1):
                    self.insertRow(row_id)
            elif insert_type == "C":
                col_id = selected.leftColumn()
                for i in range(selected.rightColumn() - selected.leftColumn() + 1):
                    self.insertColumn(col_id)
            self.selected_col = None
            self.selected_row = None
            self.renumber_header()
            self._max_content_pos()

        return insert

    def insert_base_on_selection(self):
        if not self._judge_rectangular_selected():
            return
        if self.selected_col is not None:
            self.insert_whole_base_on_selection(insert_type="C")
        elif self.selected_row is not None:
            self.insert_whole_base_on_selection(insert_type="R")
        else:
            insert_dialog = DeleteInsertDialog(dialog_type="insert")
            insert_dialog.InsertSignal.connect(self._insert_operations)
            insert_dialog.setWindowModality(Qt.WindowModality.ApplicationModal)
            insert_dialog.exec()

    @Slot(str)
    def _insert_operations(self, message):
        selected = self.selectedRanges()[0]
        if message == "Move Right":
            selected_cols_num = selected.rightColumn() - selected.leftColumn() + 1
            final_col = (
                self._insert_add_col_helper(selected.topRow(), selected.bottomRow())
                + 1
                + selected_cols_num
            )
            while self.columnCount() < final_col:
                self.insertColumn(self.columnCount())
            for col in sorted(range(selected.leftColumn(), final_col), reverse=True):
                ori_col = col - selected_cols_num
                for row in range(selected.topRow(), selected.bottomRow() + 1):
                    if col >= selected.leftColumn() + selected_cols_num:
                        text = (
                            self.item(row, ori_col).text()
                            if self.item(row, ori_col) is not None
                            else ""
                        )
                        self.setItem(row, col, CTWItem(text))
                    else:
                        self.setItem(row, col, CTWItem(""))
            self.renumber_header()
            self._max_content_pos()
        elif message == "Move Down":
            selected_rows_num = selected.bottomRow() - selected.topRow() + 1
            final_row = (
                self._insert_add_row_helper(
                    selected.leftColumn(), selected.rightColumn()
                )
                + 1
                + selected_rows_num
            )
            while self.rowCount() < final_row:
                self.insertRow(self.rowCount())
            for row in sorted(range(selected.topRow(), final_row), reverse=True):
                ori_row = row - selected_rows_num
                for col in range(selected.leftColumn(), selected.rightColumn() + 1):
                    if row >= selected.topRow() + selected_rows_num:
                        text = (
                            self.item(ori_row, col).text()
                            if self.item(ori_row, col) is not None
                            else ""
                        )
                        self.setItem(row, col, CTWItem(text))
                    else:
                        self.setItem(row, col, CTWItem(""))
            self.renumber_header()
            self._max_content_pos()
        elif message == "Insert Rows Above":
            self.insert_whole_base_on_selection("R")()
        elif message == "Insert Cols Left":
            self.insert_whole_base_on_selection("C")()
        else:
            qDebug("Empty Message")
        self.renumber_header()
        self._max_content_pos()

    def _set_null_item(self):
        if not self._judge_rectangular_selected():
            return
        s = self.selectedRanges()[0]
        for row in range(s.topRow(), s.bottomRow() + 1):
            for col in range(s.leftColumn(), s.rightColumn() + 1):
                item = self.item(row, col)
                if item is None:
                    self.setItem(row, col, CTWItem(""))

    def align_right(self):
        """
        右对齐选中区域内容
        """
        self._set_null_item()
        for item in self.selectedItems():
            item.setTextAlignment(
                Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter
            )

    def align_left(self):
        """
        左对齐选中区域内容
        """
        self._set_null_item()
        for item in self.selectedItems():
            item.setTextAlignment(
                Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter
            )

    def align_center(self):
        """
        居中对齐选中区域内容
        """
        self._set_null_item()
        for item in self.selectedItems():
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

    def highlight_background(self, choice):
        """
        高亮选中区域背景
        """

        def highlight_operation():
            self._set_null_item()
            yellow_bg = QColor(255, 255, 0)
            transparent_bg = QColor(255, 255, 255, 0)
            for item in self.selectedItems():
                if choice == "Y":
                    item.setBackground(yellow_bg)
                elif choice == "T":
                    item.setBackground(transparent_bg)
                else:
                    qDebug("highlight background error input")

        return highlight_operation

    def bold_text(self, choice):
        """
        加粗选中区域内容
        """

        def bold_operation():
            self._set_null_item()
            for item in self.selectedItems():
                item_font = item.font()
                if choice == "Y":
                    item_font.setBold(True)
                    item.setFont(item_font)
                elif choice == "N":
                    item_font.setBold(False)
                    item.setFont(item_font)
                else:
                    qDebug("bold operation error input")

        return bold_operation

    @staticmethod
    def write_dim2_list_to_table(dim2_list, tablewidget_object):
        for row, dim1_list in enumerate(dim2_list):
            for col, text in enumerate(dim1_list):
                item = CTWItem(str(text))
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                tablewidget_object.setItem(row, col, item)

    def filter_operation(self, col, filter_list):
        for row in range(self.rowCount()):
            item = self.item(row, col)
            if item is not None and item.text() in filter_list:
                continue
            else:
                self.setRowHidden(row, True)

    def _dialog_choose_sheet(
        self, sheetnames: list[str], active_names: str = None, active_index: int = 0
    ) -> tuple[str, int]:
        """
        弹出对话框选择工作表
        Args:
            sheetnames (list[str]): 工作表名称列表
            active_names (str, optional): 活动工作表名称. Defaults to None.
            active_index (int, optional): 活动工作表索引. Defaults to 0.
        Returns:
            tuple[str, int]: 工作表名称, 工作表索引
        """
        if len(sheetnames) <= 1:
            return
        if active_names in sheetnames:
            active_index = sheetnames.index(active_names)
        else:
            active_index = 0
        dialog = QDialog(self)
        dialog.setWindowTitle(self.tr("选择工作表"))
        layout = QVBoxLayout()
        label = QLabel(self.tr("请选择要加载的工作表:"))
        layout.addWidget(label)
        # Create list widget for sheets
        sheet_list = QListWidget()
        sheet_list.addItems(sheetnames)
        sheet_list.setCurrentRow(active_index)
        layout.addWidget(sheet_list)
        # Add buttons
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)
        dialog.setLayout(layout)
        if dialog.exec() != QDialog.Accepted:
            return
        sheet_name = sheet_list.currentItem().text()
        sheet_number = sheetnames.index(sheet_name)
        return sheet_name, sheet_number

    def load_csv(self, path: str):
        """
        加载CSV文件
        Args:
            path (str): 文件路径
        """
        if not path:
            return
        try:
            with open(path, "r", newline="", encoding="utf-8") as f:
                reader = csv.reader(f)
                rows = [row for row in reader if row]
                if len(rows) < 1:
                    QMessageBox.critical(
                        self, self.tr("加载错误"), self.tr("CSV文件没有内容")
                    )
                    return
                col_num = len(rows[0])
                row_num = len(rows)
                for row in range(row_num):
                    if len(rows[row]) > col_num:
                        col_num = len(rows[row])
                if self.columnCount() < col_num:
                    self.setColumnCount(col_num)
                    self.renumber_header_col()
                if self.rowCount() < row_num:
                    self.setRowCount(row_num)
                    self.renumber_header_row()
                self.clearContents()
                for row in range(row_num):
                    for col in range(col_num):
                        if rows[row][col] is None:
                            continue
                        text = str(rows[row][col])
                        if text in ["", " "]:
                            continue
                        item = self.item(row, col)
                        if item:
                            item.setText(text)
                        else:
                            self.setItem(row, col, CTWItem(text))
            qDebug("Loaded file: %s" % path)
        except Exception as e:
            QMessageBox.critical(
                self, self.tr("加载错误"), self.tr("加载文件失败\n\n%s" % e)
            )
            qDebug("Error loading file: %s" % e)

    def load_excel03(self, path: str):
        """
        加载Excel97-03文件
        Args:
            path (str): 文件路径
        """
        if not path:
            return
        try:
            book = xlrd.open_workbook(path)
            if len(book.sheet_names()) > 1:
                _, sheet_index = self._dialog_choose_sheet(
                    book.sheet_names(), active_index=0
                )
                sheet = book.sheet_by_index(sheet_index)
            else:
                sheet = book.sheet_by_index(0)
            is_empty = sheet.nrows == 0 or (
                sheet.nrows == 1 and sheet.ncols == 1 and sheet.cell_value(0, 0) == ""
            )
            if is_empty:
                QMessageBox.critical(
                    self, self.tr("加载错误"), self.tr("Excel文件没有内容")
                )
                return
            self.clearContents()
            if self.columnCount() < sheet.ncols:
                self.setColumnCount(sheet.ncols)
                self.renumber_header_col()
            if self.rowCount() < sheet.nrows:
                self.setRowCount(sheet.nrows)
                self.renumber_header_row()
            for row in range(sheet.nrows):
                r_list = sheet.row_values(row)
                for col in range(sheet.ncols):
                    text = str(r_list[col])
                    if text in ["", " "]:
                        continue
                    self.setItem(row, col, CTWItem(text))
            qDebug("Loaded file: %s" % path)
            qDebug("Loaded sheet: %s" % sheet.name)
        except Exception as e:
            QMessageBox.critical(
                self, self.tr("加载错误"), self.tr("加载文件失败\n\n%s" % e)
            )
            qDebug("Error loading file: %s" % e)

    def load_excel(self, path: str):
        """
        加载Excel文件
        Args:
            path (str): 文件路径
        """
        if not path:
            return
        try:
            book = openpyxl.load_workbook(path, data_only=True)
            if len(book.sheetnames) > 1:
                active, _ = self._dialog_choose_sheet(
                    book.sheetnames, active_names=book.active.title
                )
                sheet = book[active]
            else:
                sheet = book.active
            rows = []
            for row in sheet.iter_rows(values_only=True):
                rows.append(row)
            if len(rows) < 1:
                QMessageBox.critical(
                    self, self.tr("加载错误"), self.tr("Excel文件没有内容")
                )
                return
            col_num = len(rows[0])
            row_num = len(rows)
            for row in range(row_num):
                if len(rows[row]) > col_num:
                    col_num = len(rows[row])
            if self.columnCount() < col_num:
                self.setColumnCount(col_num)
                self.renumber_header_col()
            if self.rowCount() < row_num:
                self.setRowCount(row_num)
                self.renumber_header_row()
            self.clearContents()
            for row in range(row_num):
                for col in range(col_num):
                    if rows[row][col] is None:
                        continue
                    text = str(rows[row][col])
                    if text in ["", " "]:
                        continue
                    item = self.item(row, col)
                    if item:
                        item.setText(text)
                    else:
                        self.setItem(row, col, CTWItem(text))
            qDebug("Loaded file: %s" % path)
            qDebug("Loaded sheet: %s" % sheet.title)
        except Exception as e:
            QMessageBox.critical(
                self, self.tr("加载错误"), self.tr("加载文件失败\n\n%s" % e)
            )
            qDebug("Error loading file: %s" % e)

    def save_csv(self, path: str):
        """
        保存CSV文件
        Args:
            path (str): 文件路径
        """
        if not path:
            return
        try:
            with open(path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                # deal row data
                for row in range(self.max_content_row + 1):
                    row_data = []
                    for col in range(self.max_content_col + 1):
                        item = self.item(row, col)
                        text = item.text() if item else ""
                        if text in ("", " ", None):
                            text = " "
                        row_data.append(text)
                    writer.writerow(row_data)
            QMessageBox.information(
                self, self.tr("保存成功"), self.tr(f"文件已保存至：{path}")
            )
            qDebug("Saved file: %s" % path)
        except Exception as e:
            QMessageBox.critical(
                self,
                self.tr("保存错误"),
                self.tr("保存文件时出现错误\n\n%s" % e),
            )
            qDebug("Error saving file: %s" % e)

    def save_excel(self, path: str):
        """
        保存Excel文件
        Args:
            path (str): 文件路径
        """
        if not path:
            return
        try:
            book = openpyxl.Workbook()
            sheet = book.active
            sheet.title = "Sheet1"
            # deal row data
            for row in range(self.max_content_row + 1):
                for col in range(self.max_content_col + 1):
                    item = self.item(row, col)
                    text = item.text() if item else ""
                    if text in ("", " ", None):
                        continue
                    sheet.cell(row=row + 1, column=col + 1, value=text)
            book.save(path)
            QMessageBox.information(
                self, self.tr("保存成功"), self.tr(f"文件已保存至：{path}")
            )
            qDebug("Saved file: %s" % path)
        except Exception as e:
            QMessageBox.critical(
                self,
                self.tr("保存错误"),
                self.tr("保存文件时出现错误\n\n%s" % e),
            )
            qDebug("Error saving file: %s" % e)

    def save_excel03(self, path: str):
        """
        保存Excel97-03文件
        Args:
            path (str): 文件路径
        """
        if not path:
            return
        try:
            book = xlwt.Workbook(encoding="utf-8")
            sheet = book.add_sheet("Sheet1")
            # deal row data
            for row in range(self.max_content_row + 1):
                for col in range(self.max_content_col + 1):
                    item = self.item(row, col)
                    text = item.text() if item else ""
                    if text in ("", " ", None):
                        continue
                    sheet.write(row, col, text)
            book.save(path)
            QMessageBox.information(
                self, self.tr("保存成功"), self.tr(f"文件已保存至：{path}")
            )
            qDebug("Saved file: %s" % path)
        except Exception as e:
            QMessageBox.critical(
                self,
                self.tr("保存错误"),
                self.tr("保存文件时出现错误\n\n%s" % e),
            )
            qDebug("Error saving file: %s" % e)


class DeleteInsertDialog(QDialog):
    DelSignal = Signal(str)
    InsertSignal = Signal(str)

    def __init__(self, dialog_type):
        super().__init__()
        self.dialog_type = dialog_type
        self.setWindowFlags(Qt.WindowType.Drawer | Qt.WindowType.WindowCloseButtonHint)
        self.dialog_typ_ch = (
            self.tr("删除") if dialog_type.lower() == "delete" else self.tr("插入")
        )
        self.setWindowTitle(self.dialog_typ_ch)
        self.resize(250, 150)
        # 1. 创建GroupBox
        self.groupBox = QGroupBox(self.dialog_typ_ch)
        self.groupBox.setFlat(True)
        # 1.1 创建RadioButton
        self.radioButtonA = QRadioButton()
        self.radioButtonB = QRadioButton()
        self.radioButtonC = QRadioButton()
        self.radioButtonD = QRadioButton()
        # 1.2 将RadioButton添加到GroupBox中
        gb_vbox = QVBoxLayout()
        gb_vbox.addWidget(self.radioButtonA)
        gb_vbox.addWidget(self.radioButtonB)
        gb_vbox.addWidget(self.radioButtonC)
        gb_vbox.addWidget(self.radioButtonD)
        self.groupBox.setLayout(gb_vbox)
        # 2. 创建确定和取消按钮
        self.buttonOK = QPushButton(self.tr("确定"))
        self.buttonCancel = QPushButton(self.tr("取消"))
        # 2.1 将按钮添加到水平布局中
        hbox = QHBoxLayout()
        hbox.addWidget(self.buttonOK)
        hbox.addWidget(self.buttonCancel)
        # 3. 将GroupBox和按钮添加到垂直布局中
        vbox = QVBoxLayout()
        vbox.addWidget(self.groupBox)
        vbox.addLayout(hbox)
        # 4. 设置对话框的布局
        self.setLayout(vbox)
        self.buttonCancel.clicked.connect(self.close)
        self._preperation()

    def _preperation(self):
        if self.dialog_type.lower() == "delete":
            # 1. 设置文本
            self.radioButtonA.setText(self.tr("右侧单元格左移(L)"))
            self.radioButtonB.setText(self.tr("下方单元格上移(U)"))
            self.radioButtonC.setText(self.tr("整行(R)"))
            self.radioButtonD.setText(self.tr("整列(C)"))

            # 2 设置热键
            QShortcut(QKeySequence("L"), self).activated.connect(
                self.radioButtonA.toggle
            )
            QShortcut(QKeySequence("U"), self).activated.connect(
                self.radioButtonB.toggle
            )
            QShortcut(QKeySequence("R"), self).activated.connect(
                self.radioButtonC.toggle
            )
            QShortcut(QKeySequence("C"), self).activated.connect(
                self.radioButtonD.toggle
            )

            self.buttonOK.clicked.connect(self.delete_button_ok_clicked)
        else:
            # 1. 设置文本
            self.radioButtonA.setText(self.tr("活动单元格右移(R)"))
            self.radioButtonB.setText(self.tr("活动单元格下移(D)"))
            self.radioButtonC.setText(self.tr("整行(R)"))
            self.radioButtonD.setText(self.tr("整列(C)"))

            # 2 设置热键
            QShortcut(QKeySequence("R"), self).activated.connect(
                self.radioButtonA.toggle
            )
            QShortcut(QKeySequence("D"), self).activated.connect(
                self.radioButtonB.toggle
            )
            QShortcut(QKeySequence("R"), self).activated.connect(
                self.radioButtonC.toggle
            )
            QShortcut(QKeySequence("C"), self).activated.connect(
                self.radioButtonD.toggle
            )

            self.buttonOK.clicked.connect(self.insert_button_ok_clicked)

    def delete_button_ok_clicked(self):
        """
        删除按钮点击事件
        """
        if self.radioButtonA.isChecked():
            self.DelSignal.emit("Move Left")
        elif self.radioButtonB.isChecked():
            self.DelSignal.emit("Move Up")
        elif self.radioButtonC.isChecked():
            self.DelSignal.emit("Delete Selected Rows")
        elif self.radioButtonD.isChecked():
            self.DelSignal.emit("Delete Selected Cols")
        else:
            self.DelSignal.emit("")
        self.close()

    def insert_button_ok_clicked(self):
        """
        插入按钮点击事件
        """
        if self.radioButtonA.isChecked():
            self.InsertSignal.emit("Move Right")
        elif self.radioButtonB.isChecked():
            self.InsertSignal.emit("Move Down")
        elif self.radioButtonC.isChecked():
            self.InsertSignal.emit("Insert Rows Above")
        elif self.radioButtonD.isChecked():
            self.InsertSignal.emit("Insert Cols Left")
        else:
            self.InsertSignal.emit("")
        self.close()


if __name__ == "__main__":
    """
    test CleverTW
    """
    import sys

    class TestUsingCleverTW(QDialog):
        def __init__(self, parent=None):
            super().__init__(parent)
            self.setWindowTitle(self.tr("Test Using CleverTableWidget"))
            self.setWindowFlags(
                Qt.WindowType.WindowMinimizeButtonHint
                | Qt.WindowType.WindowMaximizeButtonHint
                | Qt.WindowType.WindowCloseButtonHint
            )
            self.tableWidget = CleverTableWidget()
            self.tableWidget.setRowCount(10)
            self.tableWidget.setColumnCount(5)
            layout = QVBoxLayout()
            layout.addWidget(self.tableWidget)
            self.setLayout(layout)
            self.add_table_content()
            self.resize(700, 400)
            self.show()

        def add_table_content(self):
            for i in range(10):
                for j in range(5):
                    item = CTWItem("{}{}".format(i + 1, j + 1))
                    self.tableWidget.setItem(i, j, item)

    app = QApplication(sys.argv)
    ui = TestUsingCleverTW()
    ui.show()
    ui.tableWidget.enable_infinite()
    ui.tableWidget.renumber_header_col()
    sys.exit(app.exec())
